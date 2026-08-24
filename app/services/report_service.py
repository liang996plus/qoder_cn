"""分期限目标检视报表 — Excel 模板填充服务"""

from __future__ import annotations

import io
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

from app.models.report_models import TermTargetReviewRequest

_TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "分期限目标检视.xlsx"

# JSON 索引 → 模板行号
_ROW_MAP = {0: 4, 1: 5, 2: 6, 3: 7, 4: 8}

_TOTAL_ROW = 9


def generate_term_target_review(request: TermTargetReviewRequest) -> bytes:
    """加载模板、填充数据、返回 xlsx bytes"""
    wb = load_workbook(_TEMPLATE_PATH)
    ws = wb.active

    _fill_header_dates(ws, request.current_date)
    _fill_data_rows(ws, request.rows)
    _fill_total_row(ws, request.rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 内部函数 ──────────────────────────────────────────


def _fill_header_dates(ws, current_date: date) -> None:
    """替换列头中的动态日期"""
    month = current_date.month
    day = current_date.day

    # C2: "{month}月规模"
    ws["C2"] = f"{month}月规模"

    # C3: "{M}/{D}规模"
    ws["C3"] = f"{month}/{day}规模"

    # E3: "主要产品销量情况（{start}-{end}）（{n}个工作日）"
    start = current_date - timedelta(days=6)
    end = current_date
    start_str = f"{start.month}/{start.day}"
    end_str = f"{end.month}/{end.day}"
    work_days = sum(
        1 for i in range(7)
        if (start + timedelta(days=i)).weekday() < 5
    )
    ws["E3"] = f"主要产品销量情况（{start_str}-{end_str}）（{work_days}个工作日）"


def _fill_data_rows(ws, rows) -> None:
    """将数据行写入对应模板行（索引制映射）"""
    for idx, row in enumerate(rows):
        template_row = _ROW_MAP[idx]
        ws.cell(row=template_row, column=3, value=round(row.daily_scale, 2))
        ws.cell(row=template_row, column=4, value=round(row.vs_last_month, 2))
        ws.cell(row=template_row, column=5, value=row.sales_analysis)


def _fill_total_row(ws, rows) -> None:
    """合计行：C/D 列求和，E 列固定 "-" """
    total_scale = round(sum(r.daily_scale for r in rows), 2)
    total_vs = round(sum(r.vs_last_month for r in rows), 2)
    ws.cell(row=_TOTAL_ROW, column=3, value=total_scale)
    ws.cell(row=_TOTAL_ROW, column=4, value=total_vs)
    # E9 已预填 "-"，保持不变
