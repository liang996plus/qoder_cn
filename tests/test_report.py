"""分期限目标检视报表 API 测试"""

import io

import openpyxl
import pytest


# ── 测试数据 ──────────────────────────────────────────

URL = "/api/v1/report/term-target-review"

FULL_PAYLOAD = {
    "current_date": "2026-06-22",
    "rows": [
        {
            "product_type": "固定收益类",
            "term_category": "超短期固收【T+1日开】",
            "daily_scale": 1200.50,
            "vs_last_month": -50.30,
            "sales_analysis": "净增长 +120.50 万元，主要产品：A、B",
        },
        {
            "product_type": "固定收益类",
            "term_category": "超短期固收【7D-2M】",
            "daily_scale": 800.00,
            "vs_last_month": 30.20,
            "sales_analysis": "净增长 -80.00 万元，主要产品：C",
        },
        {
            "product_type": "固定收益类",
            "term_category": "中短期固收",
            "daily_scale": 500.00,
            "vs_last_month": 10.00,
            "sales_analysis": "无数据",
        },
        {
            "product_type": "固定收益类",
            "term_category": "长期限固收",
            "daily_scale": 300.00,
            "vs_last_month": -5.00,
            "sales_analysis": "净增长 +50.00 万元，主要产品：D",
        },
        {
            "product_type": "含权益类",
            "term_category": "含权益类",
            "daily_scale": 200.00,
            "vs_last_month": 15.00,
            "sales_analysis": "净增长 +30.00 万元，主要产品：E",
        },
    ],
}


def _parse_xlsx(data: bytes) -> openpyxl.Workbook:
    """从 bytes 解析 openpyxl workbook"""
    return openpyxl.load_workbook(io.BytesIO(data))


# ── 基本功能 ──────────────────────────────────────────


class TestBasicReport:
    """基本功能测试"""

    def test_returns_200(self, client):
        resp = client.post(URL, json=FULL_PAYLOAD)
        assert resp.status_code == 200
        body = resp.json()
        assert body["code"] == 0

    def test_response_has_download_url(self, client):
        resp = client.post(URL, json=FULL_PAYLOAD)
        data = resp.json()["data"]
        assert "download_url" in data
        assert "/api/v1/file/download/" in data["download_url"]
        # 返回完整可点击 URL
        assert data["download_url"].startswith("http")
        assert data["format"] == "xlsx"
        assert data["column_count"] == 6
        assert data["row_count"] == 5

    def test_downloaded_file_is_valid_xlsx(self, client):
        resp = client.post(URL, json=FULL_PAYLOAD)
        filename = resp.json()["data"]["filename"]
        dl = client.get(f"/api/v1/file/download/{filename}")
        assert dl.status_code == 200
        wb = _parse_xlsx(dl.content)
        ws = wb.active
        assert ws.max_column >= 6  # 模板可能含空列


# ── Excel 内容正确性 ──────────────────────────────────


class TestExcelContent:
    """验证 Excel 单元格内容"""

    @pytest.fixture(autouse=True)
    def _setup(self, client):
        resp = client.post(URL, json=FULL_PAYLOAD)
        filename = resp.json()["data"]["filename"]
        dl = client.get(f"/api/v1/file/download/{filename}")
        self.ws = _parse_xlsx(dl.content).active

    # 标题行
    def test_title_row(self):
        assert self.ws["A1"].value == "二、分期限目标检视"

    def test_title_merged(self):
        merged = [str(m) for m in self.ws.merged_cells.ranges]
        assert "A1:F1" in merged

    # 动态列头
    def test_c2_month(self):
        assert self.ws["C2"].value == "6月规模"

    def test_c3_date(self):
        assert self.ws["C3"].value == "6/22规模"

    def test_e3_period(self):
        val = self.ws["E3"].value
        assert "6/16-6/22" in val
        assert "5个工作日" in val

    # 数据行（索引制映射）
    def test_row4_data(self):
        assert self.ws.cell(row=4, column=3).value == 1200.50
        assert self.ws.cell(row=4, column=4).value == -50.30
        assert "主要产品：A" in self.ws.cell(row=4, column=5).value

    def test_row5_data(self):
        assert self.ws.cell(row=5, column=3).value == 800.00
        assert self.ws.cell(row=5, column=4).value == 30.20

    def test_row7_data(self):
        assert self.ws.cell(row=7, column=3).value == 300.00
        assert self.ws.cell(row=7, column=4).value == -5.00

    def test_row8_equity(self):
        assert self.ws.cell(row=8, column=3).value == 200.00
        assert self.ws.cell(row=8, column=4).value == 15.00
        assert "主要产品：E" in self.ws.cell(row=8, column=5).value

    # 合计行
    def test_total_row_scale(self):
        expected = round(1200.50 + 800.00 + 500.00 + 300.00 + 200.00, 2)
        assert self.ws.cell(row=9, column=3).value == expected

    def test_total_row_vs(self):
        expected = round(-50.30 + 30.20 + 10.00 + (-5.00) + 15.00, 2)
        assert self.ws.cell(row=9, column=4).value == expected

    def test_total_row_sales(self):
        assert self.ws.cell(row=9, column=5).value == "-"

    # 合并单元格保留
    def test_fixed_income_merged(self):
        merged = [str(m) for m in self.ws.merged_cells.ranges]
        assert "A4:A7" in merged


# ── 动态日期 ──────────────────────────────────────────


class TestDynamicDates:
    """不同日期输入的动态列头测试"""

    def test_july_date(self, client):
        payload = {**FULL_PAYLOAD, "current_date": "2026-07-15"}
        resp = client.post(URL, json=payload)
        filename = resp.json()["data"]["filename"]
        dl = client.get(f"/api/v1/file/download/{filename}")
        ws = _parse_xlsx(dl.content).active
        assert ws["C2"].value == "7月规模"
        assert ws["C3"].value == "7/15规模"
        # 7/15 (Wed) - 6 = 7/9 (Thu), 工作日: Thu,Fri,Mon,Tue,Wed = 5
        assert "7/9-7/15" in ws["E3"].value
        assert "5个工作日" in ws["E3"].value

    def test_weekend_date(self, client):
        """测试周日输入"""
        payload = {**FULL_PAYLOAD, "current_date": "2026-06-21"}
        resp = client.post(URL, json=payload)
        filename = resp.json()["data"]["filename"]
        dl = client.get(f"/api/v1/file/download/{filename}")
        ws = _parse_xlsx(dl.content).active
        assert ws["C2"].value == "6月规模"
        assert ws["C3"].value == "6/21规模"
        # 6/21 (Sun) - 6 = 6/15 (Mon), Mon-Sun: 5个工作日
        assert "6/15-6/21" in ws["E3"].value
        assert "5个工作日" in ws["E3"].value


# ── 参数校验 ──────────────────────────────────────────


class TestValidation:
    """参数校验测试"""

    def test_missing_current_date(self, client):
        payload = {"rows": FULL_PAYLOAD["rows"]}
        resp = client.post(URL, json=payload)
        assert resp.status_code == 422

    def test_missing_rows(self, client):
        payload = {"current_date": "2026-06-22"}
        resp = client.post(URL, json=payload)
        assert resp.status_code == 422

    def test_empty_rows(self, client):
        payload = {"current_date": "2026-06-22", "rows": []}
        resp = client.post(URL, json=payload)
        assert resp.status_code == 422

    def test_invalid_date_format(self, client):
        payload = {"current_date": "not-a-date", "rows": FULL_PAYLOAD["rows"]}
        resp = client.post(URL, json=payload)
        assert resp.status_code == 422

    def test_too_many_rows(self, client):
        payload = {**FULL_PAYLOAD, "rows": FULL_PAYLOAD["rows"] + [FULL_PAYLOAD["rows"][0]]}
        resp = client.post(URL, json=payload)
        assert resp.status_code == 422

    def test_missing_required_field(self, client):
        row = {"product_type": "X", "term_category": "Y"}  # 缺少 daily_scale 等
        payload = {"current_date": "2026-06-22", "rows": [row]}
        resp = client.post(URL, json=payload)
        assert resp.status_code == 422


# ── 部分填充 ──────────────────────────────────────────


class TestPartialFill:
    """部分填充测试"""

    def test_3_rows_partial(self, client):
        """3 行数据：填充第 4、5、6 行，第 7、8 行保持原样"""
        payload = {
            "current_date": "2026-06-22",
            "rows": FULL_PAYLOAD["rows"][:3],
        }
        resp = client.post(URL, json=payload)
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["row_count"] == 3

        filename = data["filename"]
        dl = client.get(f"/api/v1/file/download/{filename}")
        ws = _parse_xlsx(dl.content).active

        # 第 4-6 行有数据
        assert ws.cell(row=4, column=3).value == 1200.50
        assert ws.cell(row=5, column=3).value == 800.00
        assert ws.cell(row=6, column=3).value == 500.00

        # 第 7-8 行保持模板原样（None 或原值）
        assert ws.cell(row=7, column=3).value is None
        assert ws.cell(row=8, column=3).value is None

        # 合计行只对已提供的 3 行求和
        expected_c = round(1200.50 + 800.00 + 500.00, 2)
        expected_d = round(-50.30 + 30.20 + 10.00, 2)
        assert ws.cell(row=9, column=3).value == expected_c
        assert ws.cell(row=9, column=4).value == expected_d

    def test_1_row_minimal(self, client):
        """仅 1 行数据"""
        payload = {
            "current_date": "2026-06-22",
            "rows": [FULL_PAYLOAD["rows"][0]],
        }
        resp = client.post(URL, json=payload)
        assert resp.status_code == 200
        filename = resp.json()["data"]["filename"]
        dl = client.get(f"/api/v1/file/download/{filename}")
        ws = _parse_xlsx(dl.content).active

        assert ws.cell(row=4, column=3).value == 1200.50
        assert ws.cell(row=9, column=3).value == 1200.50
        assert ws.cell(row=9, column=4).value == -50.30
